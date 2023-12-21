from serial.tools import list_ports

import pydobot
import openpyxl

def excel_report(p):
    wb = openpyxl.load_workbook('pydobot_pose.xlsx')
    ws = wb["Sheet1"]
    row=2
    col=1
    while True:
        if(ws.cell(row,col).value is None):
            ws.cell(row,col).value=row-1
            col+=1
            for i in p:
                ws.cell(row,col).value=i
                col+=1
            break
        row+=1
    wb.save('pydobot_pose.xlsx')

def main():
    available_ports = list_ports.comports()
    print(f'available ports: {[x.device for x in available_ports]}')
    port = available_ports[0].device

    device = pydobot.Dobot(port=port, verbose=True)

    (x, y, z, r, j1, j2, j3, j4) = device.pose()
    print(f'x:{x} y:{y} z:{z} r:{r} j1:{j1} j2:{j2} j3:{j3} j4:{j4}')

    dobot_pose=[x,y,z,r]

    excel_report(dobot_pose)
    device.close()

main()

